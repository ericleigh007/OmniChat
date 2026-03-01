"""Tests for save_output.py — file saving utilities."""

import re
from pathlib import Path

import pytest

import tools.output.save_output as so


@pytest.fixture(autouse=True)
def patch_outputs_dir(tmp_outputs_dir, monkeypatch):
    """Redirect OUTPUTS_DIR to temp directory for all tests."""
    monkeypatch.setattr(so, "OUTPUTS_DIR", tmp_outputs_dir)


class TestTimestampStem:
    """Test timestamp-based filename generation."""

    def test_default_prefix(self):
        stem = so._timestamp_stem()
        assert stem.startswith("output_")
        # Pattern: output_YYYY-MM-DD_HHMMSS
        assert re.match(r"output_\d{4}-\d{2}-\d{2}_\d{6}", stem)

    def test_custom_prefix(self):
        stem = so._timestamp_stem("scan")
        assert stem.startswith("scan_")


class TestUniquePath:
    """Test unique path generation with collision avoidance."""

    def test_no_collision(self, tmp_outputs_dir):
        path = so._unique_path(tmp_outputs_dir, "test", ".txt")
        assert path == tmp_outputs_dir / "test.txt"

    def test_collision_increments(self, tmp_outputs_dir):
        # Create the first file
        (tmp_outputs_dir / "test.txt").write_text("first")
        path = so._unique_path(tmp_outputs_dir, "test", ".txt")
        assert path == tmp_outputs_dir / "test_2.txt"

    def test_multiple_collisions(self, tmp_outputs_dir):
        (tmp_outputs_dir / "test.txt").write_text("first")
        (tmp_outputs_dir / "test_2.txt").write_text("second")
        path = so._unique_path(tmp_outputs_dir, "test", ".txt")
        assert path == tmp_outputs_dir / "test_3.txt"


class TestSaveAsMarkdown:
    """Test markdown file saving."""

    def test_creates_md_file(self, tmp_outputs_dir):
        path = so.save_as_markdown("# Hello World", filename="test_doc")
        assert Path(path).exists()
        assert path.endswith(".md")
        assert Path(path).read_text(encoding="utf-8") == "# Hello World"

    def test_auto_filename(self, tmp_outputs_dir):
        path = so.save_as_markdown("content")
        assert Path(path).exists()
        assert "scan_" in Path(path).stem


class TestSaveAsText:
    """Test plain text file saving."""

    def test_creates_txt_file(self, tmp_outputs_dir):
        path = so.save_as_text("Hello World", filename="test_text")
        assert Path(path).exists()
        assert path.endswith(".txt")
        assert Path(path).read_text(encoding="utf-8") == "Hello World"


class TestSaveAsExcel:
    """Test Excel file saving."""

    def test_creates_xlsx_file(self, tmp_outputs_dir):
        data = [["Name", "Age"], ["Alice", "30"], ["Bob", "25"]]
        path = so.save_as_excel(data, filename="test_table")
        assert Path(path).exists()
        assert path.endswith(".xlsx")

    def test_header_row_bold(self, tmp_outputs_dir):
        from openpyxl import load_workbook

        data = [["Name", "Age"], ["Alice", "30"]]
        path = so.save_as_excel(data, filename="bold_test")
        wb = load_workbook(path)
        ws = wb.active
        assert ws.cell(row=1, column=1).font.bold is True
        assert ws.cell(row=2, column=1).font.bold is not True

    def test_correct_cell_values(self, tmp_outputs_dir):
        from openpyxl import load_workbook

        data = [["A", "B"], ["1", "2"]]
        path = so.save_as_excel(data, filename="values_test")
        wb = load_workbook(path)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "A"
        assert ws.cell(row=2, column=2).value == "2"


class TestSaveAsCsv:
    """Test CSV file saving."""

    def test_creates_csv_file(self, tmp_outputs_dir):
        data = [["Name", "Age"], ["Alice", "30"], ["Bob", "25"]]
        path = so.save_as_csv(data, filename="test_csv")
        assert Path(path).exists()
        assert path.endswith(".csv")

    def test_csv_content(self, tmp_outputs_dir):
        data = [["A", "B"], ["1", "2"]]
        path = so.save_as_csv(data, filename="content_test")
        content = Path(path).read_text(encoding="utf-8")
        assert "A,B" in content
        assert "1,2" in content

    def test_csv_handles_commas_in_values(self, tmp_outputs_dir):
        import csv

        data = [["Name", "Note"], ["Alice", "Hello, world"]]
        path = so.save_as_csv(data, filename="comma_test")
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert rows[1][1] == "Hello, world"

    def test_auto_filename(self, tmp_outputs_dir):
        data = [["A"], ["1"]]
        path = so.save_as_csv(data)
        assert Path(path).exists()
        assert "table_" in Path(path).stem


class TestSaveAuto:
    """Test auto-format selection and saving."""

    def test_auto_detects_markdown(self, tmp_outputs_dir):
        content = "# Title\n\n## Section\n\n- Item one\n- Item two"
        path = so.save_auto(content, fmt="auto")
        assert path.endswith(".md")

    def test_auto_detects_text(self, tmp_outputs_dir):
        content = "Just a plain sentence."
        path = so.save_auto(content, fmt="auto")
        assert path.endswith(".txt")

    def test_explicit_markdown(self, tmp_outputs_dir):
        path = so.save_auto("plain content", fmt="markdown")
        assert path.endswith(".md")

    def test_explicit_text(self, tmp_outputs_dir):
        path = so.save_auto("# looks like md", fmt="text")
        assert path.endswith(".txt")

    def test_explicit_excel_with_table(self, tmp_outputs_dir):
        table = [["A", "B"], ["1", "2"]]
        path = so.save_auto("| A | B |", fmt="excel", table=table)
        assert path.endswith(".xlsx")

    def test_explicit_csv_with_table(self, tmp_outputs_dir):
        table = [["A", "B"], ["1", "2"]]
        path = so.save_auto("| A | B |", fmt="csv", table=table)
        assert path.endswith(".csv")

    def test_csv_without_table_falls_to_text(self, tmp_outputs_dir):
        path = so.save_auto("just text", fmt="csv", table=None)
        assert path.endswith(".txt")

    def test_explicit_tsv_with_table(self, tmp_outputs_dir):
        table = [["A", "B"], ["1", "2"]]
        path = so.save_auto("| A | B |", fmt="tsv", table=table)
        assert path.endswith(".tsv")

    def test_tsv_without_table_falls_to_text(self, tmp_outputs_dir):
        path = so.save_auto("just text", fmt="tsv", table=None)
        assert path.endswith(".txt")


class TestTsvSupport:
    """Test tab-separated file saving."""

    def test_creates_tsv_file(self, tmp_outputs_dir):
        data = [["Name", "Age"], ["Alice", "30"]]
        path = so.save_as_csv(data, filename="test_tsv", delimiter="\t")
        assert Path(path).exists()
        assert path.endswith(".tsv")

    def test_tsv_content_tab_delimited(self, tmp_outputs_dir):
        data = [["A", "B"], ["1", "2"]]
        path = so.save_as_csv(data, filename="tab_test", delimiter="\t")
        content = Path(path).read_text(encoding="utf-8")
        assert "A\tB" in content
        assert "1\t2" in content

    def test_tsv_commas_not_escaped(self, tmp_outputs_dir):
        data = [["Name", "Note"], ["Alice", "Hello, world"]]
        path = so.save_as_csv(data, filename="comma_in_tsv", delimiter="\t")
        content = Path(path).read_text(encoding="utf-8")
        # Comma is just a regular character in TSV — no quoting needed
        assert "Hello, world" in content


class TestExplicitPath:
    """Test saving to a user-specified path (Save As... dialog)."""

    def test_csv_to_explicit_path(self, tmp_path):
        data = [["A", "B"], ["1", "2"]]
        out = str(tmp_path / "my_output.csv")
        path = so.save_as_csv(data, path=out)
        assert path == out
        assert Path(path).exists()

    def test_tsv_to_explicit_path(self, tmp_path):
        data = [["A", "B"], ["1", "2"]]
        out = str(tmp_path / "my_output.tsv")
        path = so.save_as_csv(data, path=out, delimiter="\t")
        assert path == out
        assert Path(path).exists()

    def test_excel_to_explicit_path(self, tmp_path):
        data = [["A", "B"], ["1", "2"]]
        out = str(tmp_path / "my_output.xlsx")
        path = so.save_as_excel(data, path=out)
        assert path == out
        assert Path(path).exists()

    def test_markdown_to_explicit_path(self, tmp_path):
        out = str(tmp_path / "my_output.md")
        path = so.save_as_markdown("# Title", path=out)
        assert path == out
        assert Path(path).read_text(encoding="utf-8") == "# Title"

    def test_text_to_explicit_path(self, tmp_path):
        out = str(tmp_path / "my_output.txt")
        path = so.save_as_text("plain text", path=out)
        assert path == out
        assert Path(path).read_text(encoding="utf-8") == "plain text"

    def test_explicit_path_creates_parent_dirs(self, tmp_path):
        out = str(tmp_path / "sub" / "dir" / "output.csv")
        data = [["A"], ["1"]]
        path = so.save_as_csv(data, path=out)
        assert Path(path).exists()


class TestAppendMode:
    """Test appending to existing files."""

    def test_csv_append_skips_header(self, tmp_path):
        import csv

        out = str(tmp_path / "data.csv")
        batch1 = [["Date", "Amount"], ["01/01", "100"]]
        batch2 = [["Date", "Amount"], ["02/01", "200"]]

        so.save_as_csv(batch1, path=out)
        so.save_as_csv(batch2, path=out, append=True)

        with open(out, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert len(rows) == 3  # 1 header + 2 data rows
        assert rows[0] == ["Date", "Amount"]
        assert rows[1] == ["01/01", "100"]
        assert rows[2] == ["02/01", "200"]

    def test_tsv_append(self, tmp_path):
        import csv

        out = str(tmp_path / "data.tsv")
        batch1 = [["A", "B"], ["1", "2"]]
        batch2 = [["A", "B"], ["3", "4"]]

        so.save_as_csv(batch1, path=out, delimiter="\t")
        so.save_as_csv(batch2, path=out, delimiter="\t", append=True)

        with open(out, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f, delimiter="\t"))
        assert len(rows) == 3
        assert rows[2] == ["3", "4"]

    def test_csv_append_to_nonexistent_writes_all(self, tmp_path):
        out = str(tmp_path / "new.csv")
        data = [["H"], ["r1"]]
        so.save_as_csv(data, path=out, append=True)
        content = Path(out).read_text(encoding="utf-8")
        # File didn't exist, so append=True still writes everything
        assert "H" in content
        assert "r1" in content

    def test_excel_append_adds_rows(self, tmp_path):
        from openpyxl import load_workbook

        out = str(tmp_path / "data.xlsx")
        batch1 = [["Name", "Age"], ["Alice", "30"]]
        batch2 = [["Name", "Age"], ["Bob", "25"]]

        so.save_as_excel(batch1, path=out)
        so.save_as_excel(batch2, path=out, append=True)

        wb = load_workbook(out)
        ws = wb.active
        assert ws.max_row == 3  # header + 2 data rows
        assert ws.cell(row=1, column=1).value == "Name"
        assert ws.cell(row=2, column=1).value == "Alice"
        assert ws.cell(row=3, column=1).value == "Bob"

    def test_excel_append_preserves_bold_header(self, tmp_path):
        from openpyxl import load_workbook

        out = str(tmp_path / "data.xlsx")
        batch1 = [["H1", "H2"], ["a", "b"]]
        batch2 = [["H1", "H2"], ["c", "d"]]

        so.save_as_excel(batch1, path=out)
        so.save_as_excel(batch2, path=out, append=True)

        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=1, column=1).font.bold is True
        assert ws.cell(row=3, column=1).font.bold is not True

    def test_text_append(self, tmp_path):
        out = str(tmp_path / "log.txt")
        so.save_as_text("First scan", path=out)
        so.save_as_text("Second scan", path=out, append=True)

        content = Path(out).read_text(encoding="utf-8")
        assert "First scan" in content
        assert "Second scan" in content
        assert content.index("First scan") < content.index("Second scan")

    def test_markdown_append(self, tmp_path):
        out = str(tmp_path / "notes.md")
        so.save_as_markdown("# Page 1", path=out)
        so.save_as_markdown("# Page 2", path=out, append=True)

        content = Path(out).read_text(encoding="utf-8")
        assert "# Page 1" in content
        assert "# Page 2" in content

    def test_three_csv_appends(self, tmp_path):
        """Simulate 3 monthly bank statement scans saved to one file."""
        import csv

        out = str(tmp_path / "yearly.csv")
        jan = [["Date", "Amt"], ["01/01", "10"], ["01/15", "20"]]
        feb = [["Date", "Amt"], ["02/01", "30"]]
        mar = [["Date", "Amt"], ["03/01", "40"], ["03/20", "50"]]

        so.save_as_csv(jan, path=out)
        so.save_as_csv(feb, path=out, append=True)
        so.save_as_csv(mar, path=out, append=True)

        with open(out, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert len(rows) == 6  # 1 header + 5 data rows
        assert rows[0] == ["Date", "Amt"]
        assert rows[-1] == ["03/20", "50"]
