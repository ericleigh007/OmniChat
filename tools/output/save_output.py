"""
save_output.py — Save processed content to markdown, text, CSV, or Excel files.

Auto-detects the best format based on content structure, or accepts
an explicit format override. All files are saved to the configured
output directory (default: outputs/).
"""

import sys
from pathlib import Path
from datetime import datetime

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

BASE_DIR = Path(__file__).parent.parent.parent
OUTPUTS_DIR = BASE_DIR / "outputs"


def _ensure_dir():
    """Create outputs directory if it doesn't exist."""
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)


def _unique_path(directory: Path, stem: str, suffix: str) -> Path:
    """Generate a unique file path, appending _2, _3, etc. if needed."""
    path = directory / f"{stem}{suffix}"
    if not path.exists():
        return path

    counter = 2
    while True:
        path = directory / f"{stem}_{counter}{suffix}"
        if not path.exists():
            return path
        counter += 1


def _timestamp_stem(prefix: str = "output") -> str:
    """Generate a filename stem with timestamp: output_2026-02-21_143052."""
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    return f"{prefix}_{ts}"


def save_as_markdown(
    content: str,
    filename: str | None = None,
    path: str | None = None,
    append: bool = False,
) -> str:
    """
    Save content as a markdown file.

    Args:
        content: Text content (may include markdown formatting).
        filename: Optional filename (without extension). Auto-generated if None.
                  Ignored when path is provided.
        path: Explicit output path. When provided, saves directly to this path
              instead of auto-generating in OUTPUTS_DIR.
        append: If True and file exists, append content instead of overwriting.

    Returns:
        Path to the saved file.
    """
    if path:
        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
    else:
        _ensure_dir()
        stem = filename or _timestamp_stem("scan")
        out = _unique_path(OUTPUTS_DIR, stem, ".md")
    if append and out.exists():
        with open(out, "a", encoding="utf-8") as f:
            f.write("\n\n" + content)
    else:
        out.write_text(content, encoding="utf-8")
    return str(out)


def save_as_text(
    content: str,
    filename: str | None = None,
    path: str | None = None,
    append: bool = False,
) -> str:
    """
    Save content as a plain text file.

    Args:
        content: Text content.
        filename: Optional filename (without extension). Auto-generated if None.
                  Ignored when path is provided.
        path: Explicit output path. When provided, saves directly to this path
              instead of auto-generating in OUTPUTS_DIR.
        append: If True and file exists, append content instead of overwriting.

    Returns:
        Path to the saved file.
    """
    if path:
        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
    else:
        _ensure_dir()
        stem = filename or _timestamp_stem("scan")
        out = _unique_path(OUTPUTS_DIR, stem, ".txt")
    if append and out.exists():
        with open(out, "a", encoding="utf-8") as f:
            f.write("\n\n" + content)
    else:
        out.write_text(content, encoding="utf-8")
    return str(out)


def save_as_excel(
    data: list[list[str]],
    filename: str | None = None,
    sheet_name: str = "Sheet1",
    path: str | None = None,
    append: bool = False,
) -> str:
    """
    Save tabular data as an Excel file.

    Args:
        data: List of rows, where each row is a list of cell strings.
              The first row is treated as headers.
        filename: Optional filename (without extension). Auto-generated if None.
                  Ignored when path is provided.
        sheet_name: Name of the Excel worksheet.
        path: Explicit output path. When provided, saves directly to this path
              instead of auto-generating in OUTPUTS_DIR.
        append: If True and file exists, append data rows (skip header) to
                the existing worksheet instead of overwriting.

    Returns:
        Path to the saved file.
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font

    if path:
        out_path = Path(path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
    else:
        _ensure_dir()
        stem = filename or _timestamp_stem("table")
        out_path = _unique_path(OUTPUTS_DIR, stem, ".xlsx")

    if append and out_path.exists():
        wb = load_workbook(str(out_path))
        ws = wb.active
        # Append data rows only (skip header row[0])
        for row in data[1:]:
            ws.append(row)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for row_idx, row in enumerate(data):
            for col_idx, cell in enumerate(row):
                ws_cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell)
                if row_idx == 0:
                    ws_cell.font = Font(bold=True)
        # Auto-fit column widths (approximate)
        for col_idx in range(len(data[0]) if data else 0):
            max_len = 0
            for row in data:
                if col_idx < len(row):
                    max_len = max(max_len, len(str(row[col_idx])))
            ws.column_dimensions[chr(65 + col_idx) if col_idx < 26 else "A"].width = min(max_len + 2, 50)

    wb.save(str(out_path))
    return str(out_path)


def save_as_csv(
    data: list[list[str]],
    filename: str | None = None,
    delimiter: str = ",",
    path: str | None = None,
    append: bool = False,
) -> str:
    """
    Save tabular data as a CSV (or TSV) file.

    Args:
        data: List of rows, where each row is a list of cell strings.
              The first row is treated as headers.
        filename: Optional filename (without extension). Auto-generated if None.
                  Ignored when path is provided.
        delimiter: Field delimiter — ',' for CSV, '\\t' for TSV.
        path: Explicit output path. When provided, saves directly to this path
              instead of auto-generating in OUTPUTS_DIR.
        append: If True and file exists, append data rows (skip header) to
                the existing file instead of overwriting.

    Returns:
        Path to the saved file.
    """
    import csv

    if path:
        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
    else:
        _ensure_dir()
        stem = filename or _timestamp_stem("table")
        ext = ".tsv" if delimiter == "\t" else ".csv"
        out = _unique_path(OUTPUTS_DIR, stem, ext)

    if append and out.exists():
        with open(out, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerows(data[1:])  # skip header
    else:
        with open(out, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerows(data)

    return str(out)


def save_auto(content: str, fmt: str = "auto", table: list[list[str]] | None = None, filename: str | None = None) -> str:
    """
    Save content in the best format, auto-detecting if needed.

    Args:
        content: The text content to save.
        fmt: Format hint — 'auto', 'markdown', 'text', 'excel', 'csv', or 'tsv'.
        table: Pre-parsed table data (list of rows). Required for excel/csv/tsv.
        filename: Optional filename (without extension).

    Returns:
        Path to the saved file.
    """
    if fmt == "auto":
        # Use the format detection from process_media
        from tools.vision.process_media import _detect_format
        fmt = _detect_format(content)

    if fmt == "excel" and table:
        return save_as_excel(table, filename=filename)
    elif fmt == "csv" and table:
        return save_as_csv(table, filename=filename)
    elif fmt == "tsv" and table:
        return save_as_csv(table, filename=filename, delimiter="\t")
    elif fmt == "markdown":
        return save_as_markdown(content, filename=filename)
    else:
        return save_as_text(content, filename=filename)


# ── CLI test ──────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Test output saving")
    parser.add_argument("--format", choices=["auto", "markdown", "text", "excel"], default="auto")
    args = parser.parse_args()

    # Test with sample content
    if args.format == "excel":
        test_data = [
            ["Name", "Age", "City"],
            ["Alice", "30", "New York"],
            ["Bob", "25", "London"],
            ["Charlie", "35", "Tokyo"],
        ]
        path = save_as_excel(test_data, filename="test_table")
        print(f"Saved Excel: {path}")

    elif args.format == "markdown":
        test_md = "# Test Document\n\n## Section 1\n\n- Item one\n- Item two\n\n## Section 2\n\nSome paragraph text here."
        path = save_as_markdown(test_md, filename="test_markdown")
        print(f"Saved Markdown: {path}")

    else:
        test_text = "This is a plain text test.\nLine two.\nLine three."
        path = save_as_text(test_text, filename="test_text")
        print(f"Saved Text: {path}")

    print("Done.")
